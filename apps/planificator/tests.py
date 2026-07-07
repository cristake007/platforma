import io
from datetime import datetime, timedelta, timezone as datetime_timezone
from unittest.mock import patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Permission
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.db import DatabaseError
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .file_handlers import create_excel_export, read_input_file
from .models import AppSetting, ScheduleGeneration
from .settings_store import get_settings, save_settings


def csv_upload(name="courses.csv", rows=None):
    rows = rows or [
        "Alpha Upload,3 days,https://example.com/course-a,100",
        "Beta Upload,2 days,https://example.com/course-b,200",
    ]
    content = "Title,Durata Curs,Permalink,investitie\n" + "\n".join(rows) + "\n"
    return SimpleUploadedFile(name, content.encode("utf-8"), content_type="text/csv")


def generation_for(user, *, schedule=None, expired=False):
    schedule = schedule or [{
        "source_row": 2,
        "original_order": 0,
        "Title": "Course A",
        "Permalink": "https://example.com/course-a",
        "Durata Curs": "3 days",
        "investitie": "100",
        "date_range": "05-07.01.2026",
        "month": 1,
        "month_name": "Ianuarie",
    }]
    return ScheduleGeneration.objects.create(
        owner=user,
        year=2026,
        selected_months=[1],
        holidays=[],
        random_seed=42,
        schedule=schedule,
        source_course_count=1,
        generated_entry_count=len(schedule),
        source_file_name="courses.csv",
        source_file_digest="a" * 64,
        source_file_data=(
            b"Title,Durata Curs,Permalink,investitie\n"
            b"Course A,3 days,https://example.com/course-a,100\n"
        ),
        expires_at=timezone.now() + (-timedelta(hours=1) if expired else timedelta(hours=24)),
    )


class PlanificatorViewTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = get_user_model().objects.create_user(
            username="planner", password="test-password", first_name="Plan", last_name="Operator"
        )
        cls.permission = Permission.objects.get(codename="use_course_planning")
        cls.user.user_permissions.add(cls.permission)

    def setUp(self):
        self.client.force_login(self.user)

    def test_planificator_routes_render_for_authorized_user(self):
        for route, template in (
            ("planificator:generator_perioade", "planificator/generator_perioade.html"),
            ("planificator:actualizeaza_cursuri", "planificator/actualizeaza_cursuri.html"),
            ("planificator:istoric", "planificator/istoric.html"),
        ):
            response = self.client.get(reverse(route))
            self.assertEqual(response.status_code, 200)
            self.assertTemplateUsed(response, template)

    def test_planificator_has_no_landing_page(self):
        response = self.client.get("/planificator/")

        self.assertEqual(response.status_code, 404)

    def test_generator_uses_daisyui_year_selector(self):
        response = self.client.get(reverse("planificator:generator_perioade"))

        self.assertContains(response, '<select name="year" class="select select-primary select-sm w-full" id="id_year">')
        self.assertContains(response, 'class="checkbox checkbox-primary checkbox-xs"', count=12)
        self.assertContains(response, 'type="date" id="ops-holiday-date"')
        self.assertContains(response, 'style="max-width:1360px"')
        self.assertContains(response, 'class="steps steps-vertical w-full text-xs font-semibold text-primary"', count=3)
        self.assertContains(response, 'class="generator-card-step min-w-0', count=3)
        self.assertContains(response, 'space-y-4 p-4', count=1)
        self.assertContains(response, 'data-current-step="1"')
        self.assertContains(response, 'step-secondary" data-generator-step="1"', count=1)
        self.assertNotContains(response, 'class="divider')
        self.assertContains(response, "<li>Planificator</li>")
        self.assertNotContains(response, 'href="/planificator/"')
        self.assertContains(response, 'id="generator-upload" class="card generator-step-card')
        self.assertContains(response, 'id="generator-settings" class="card generator-step-card')
        self.assertContains(response, 'id="generator-result" class="card generator-step-card')
        self.assertContains(response, 'data-generator-card="', count=3)
        self.assertContains(response, 'class="grid gap-4 md:grid-cols-3 md:items-start"')
        self.assertContains(response, 'id="ops-upload-dropzone"')
        self.assertContains(response, 'id="ops-upload-warning" class="alert alert-warning')
        self.assertFalse(response.context["form"].fields["input_file"].required)
        self.assertContains(response, "Programul este pregătit pentru generare")
        self.assertNotContains(response, "btn-xs")
        self.assertNotContains(response, 'class="btn btn-primary')
        self.assertNotContains(response, "ops-work-section")

    def test_result_table_has_its_own_scroll_area_and_three_sticky_columns(self):
        generation = generation_for(self.user)
        response = self.client.get(
            reverse("planificator:generator_perioade_result", kwargs={"generation_id": generation.pk})
        )

        self.assertContains(response, 'id="ops-preview-table-wrap" class="max-h-[32rem] overflow-auto')
        self.assertContains(response, "ops-schedule-table", count=1)
        self.assertContains(response, "ops-schedule-course-column", count=2)
        self.assertContains(response, "ops-schedule-duration-column", count=2)
        self.assertContains(response, "ops-schedule-investment-column", count=2)
        self.assertNotContains(response, "position:sticky!important")
        self.assertContains(response, 'data-current-step="3"')
        self.assertContains(response, 'step-secondary" data-generator-step="3"', count=1)
        self.assertContains(response, 'id="generator-result" class="card generator-step-card overflow-hidden')
        self.assertContains(response, f'value="{generation.pk}"')
        self.assertContains(response, "Fișier procesat · 1 cursuri")
        self.assertContains(response, "Alege alt fișier")
        self.assertContains(response, 'id="ops-clear-processed-upload"')
        self.assertContains(response, "Șterge fișierul")
        self.assertNotContains(response, "Programul a fost generat")

    def test_anonymous_user_is_redirected_to_login(self):
        self.client.logout()
        response = self.client.get(reverse("planificator:generator_perioade"))
        self.assertRedirects(response, f"{reverse('login')}?next={reverse('planificator:generator_perioade')}")

    def test_authenticated_user_without_permission_receives_403(self):
        other = get_user_model().objects.create_user(username="no-access", password="x")
        self.client.force_login(other)
        self.assertEqual(self.client.get(reverse("planificator:generator_perioade")).status_code, 403)

    def test_staff_without_permission_receives_403_and_superuser_is_allowed(self):
        staff = get_user_model().objects.create_user(username="staff", password="x", is_staff=True)
        self.client.force_login(staff)
        self.assertEqual(self.client.get(reverse("planificator:generator_perioade")).status_code, 403)
        superuser = get_user_model().objects.create_superuser(username="root-user", password="x")
        self.client.force_login(superuser)
        self.assertEqual(self.client.get(reverse("planificator:generator_perioade")).status_code, 200)

    def test_generator_uses_post_redirect_get_and_complete_matrix(self):
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"input_file": csv_upload(), "year": 2026, "months": ["1", "2"], "randomness": 5, "holidays": ""},
        )
        generation = ScheduleGeneration.objects.get()
        self.assertRedirects(
            response,
            reverse("planificator:generator_perioade_result", kwargs={"generation_id": generation.pk}),
        )
        self.assertEqual(generation.source_course_count, 2)
        self.assertEqual(generation.generated_entry_count, 4)
        self.assertEqual(
            {(row["source_row"], row["month"]) for row in generation.schedule},
            {(2, 1), (2, 2), (3, 1), (3, 2)},
        )
        self.assertTrue(generation.source_file_data)

    def test_htmx_generator_validation_error_returns_workflow_partial(self):
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"year": 2026, "months": ["1"], "randomness": 5, "holidays": ""},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "planificator/includes/generator_workflow.html")
        self.assertContains(response, 'id="generator-workflow"')
        self.assertContains(response, "Formularul este incomplet")
        self.assertContains(response, "CSV sau XLSX")
        self.assertNotContains(response, "<html")

    def test_htmx_generator_success_returns_result_workflow_partial(self):
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"input_file": csv_upload(), "year": 2026, "months": ["1", "2"], "randomness": 5, "holidays": ""},
            HTTP_HX_REQUEST="true",
        )
        generation = ScheduleGeneration.objects.get()

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "planificator/includes/generator_workflow.html")
        self.assertContains(response, 'id="generator-workflow"')
        self.assertContains(response, 'id="generator-result" class="card generator-step-card overflow-hidden')
        self.assertContains(response, "ops-schedule-table", count=1)
        self.assertContains(response, 'id="export-form"')
        self.assertContains(response, f'value="{generation.pk}"')
        self.assertContains(response, "Program generat")
        self.assertNotContains(response, "<html")

    def test_result_can_regenerate_with_saved_upload_after_year_change(self):
        initial_response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"input_file": csv_upload(), "year": 2026, "months": ["1"], "randomness": 5, "holidays": ""},
        )
        initial_generation = ScheduleGeneration.objects.get()
        self.assertEqual(initial_response.status_code, 302)

        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {
                "source_generation_id": initial_generation.pk,
                "year": 2027,
                "months": ["1"],
                "randomness": 5,
                "holidays": "",
            },
        )

        self.assertEqual(response.status_code, 302)
        regenerated = ScheduleGeneration.objects.exclude(pk=initial_generation.pk).get()
        self.assertEqual(regenerated.year, 2027)
        self.assertEqual(regenerated.source_file_digest, initial_generation.source_file_digest)
        self.assertEqual(bytes(regenerated.source_file_data), bytes(initial_generation.source_file_data))

    def test_saved_upload_regeneration_rejects_foreign_and_expired_sources(self):
        other = get_user_model().objects.create_user(username="regeneration-other")
        foreign = generation_for(other)
        expired = generation_for(self.user, expired=True)

        for generation in (foreign, expired):
            with self.subTest(generation=generation.pk):
                response = self.client.post(
                    reverse("planificator:generator_perioade"),
                    {
                        "source_generation_id": generation.pk,
                        "year": 2027,
                        "months": ["1"],
                        "randomness": 5,
                        "holidays": "",
                    },
                )

                self.assertEqual(response.status_code, 404)

    def test_missing_upload_uses_inline_warning_instead_of_required_file_input(self):
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"year": 2026, "months": ["1"], "randomness": 5, "holidays": ""},
        )

        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "alert alert-warning", status_code=400)
        self.assertContains(response, "Selectează un fișier CSV sau XLSX", status_code=400)

    def test_result_export_form_contains_only_generation_identifier(self):
        generation = generation_for(self.user)
        response = self.client.get(
            reverse("planificator:generator_perioade_result", kwargs={"generation_id": generation.pk})
        )
        self.assertContains(response, 'id="export-form"')
        self.assertContains(response, 'name="generation_id"')
        self.assertNotContains(response, "schedule_payload")

    def test_history_lists_only_owned_active_generations_with_download_actions(self):
        owned = generation_for(self.user)
        owned.source_file_name = "owned.csv"
        owned.save(update_fields=["source_file_name"])
        expired = generation_for(self.user, expired=True)
        expired.source_file_name = "expired.csv"
        expired.save(update_fields=["source_file_name"])
        other = get_user_model().objects.create_user(username="history-other")
        foreign = generation_for(other)
        foreign.source_file_name = "foreign.csv"
        foreign.save(update_fields=["source_file_name"])

        response = self.client.get(reverse("planificator:istoric"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "planificator/istoric.html")
        self.assertContains(response, "owned.csv")
        self.assertNotContains(response, "expired.csv")
        self.assertNotContains(response, "foreign.csv")
        self.assertContains(response, f'name="generation_id" value="{owned.pk}"')
        self.assertContains(response, reverse("planificator:generator_perioade_export"))
        self.assertContains(response, reverse("planificator:istoric_detail", kwargs={"generation_id": owned.pk}))
        self.assertContains(response, "Descarcă XLSX")

    def test_htmx_history_returns_list_partial(self):
        owned = generation_for(self.user)
        owned.source_file_name = "owned.csv"
        owned.save(update_fields=["source_file_name"])

        response = self.client.get(reverse("planificator:istoric"), HTTP_HX_REQUEST="true")

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "planificator/includes/history_list.html")
        self.assertContains(response, 'id="history-list"')
        self.assertContains(response, "owned.csv")
        self.assertContains(response, 'hx-get="?page=1"')
        self.assertContains(response, reverse("planificator:generator_perioade_export"))
        self.assertNotContains(response, "<html")

    def test_history_displays_generation_time_in_bucharest(self):
        generation = generation_for(self.user)
        ScheduleGeneration.objects.filter(pk=generation.pk).update(
            created_at=datetime(2026, 7, 3, 8, 7, tzinfo=datetime_timezone.utc),
            expires_at=datetime(2099, 7, 4, 8, 7, tzinfo=datetime_timezone.utc),
        )

        response = self.client.get(reverse("planificator:istoric"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "03.07.2026 11:07", count=2)

    def test_history_has_empty_state(self):
        response = self.client.get(reverse("planificator:istoric"))

        self.assertContains(response, "Nu există programe disponibile")
        self.assertContains(response, "Deschide generatorul")

    def test_history_detail_is_read_only_but_keeps_export_available(self):
        generation = generation_for(self.user)

        response = self.client.get(
            reverse("planificator:istoric_detail", kwargs={"generation_id": generation.pk})
        )

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "planificator/generator_perioade.html")
        self.assertTrue(response.context["history_read_only"])
        self.assertContains(response, 'class="contents" disabled aria-disabled="true"')
        self.assertContains(response, "Fișierul sursă și setările nu pot fi modificate")
        self.assertNotContains(response, "Alege alt fișier")
        self.assertNotContains(response, "Șterge fișierul")
        self.assertNotContains(response, "Descarcă modelul CSV")
        self.assertContains(response, 'id="export-form"')
        self.assertContains(response, "Descarcă XLSX")

    def test_history_detail_cannot_open_another_users_generation(self):
        other = get_user_model().objects.create_user(username="history-detail-other")
        generation = generation_for(other)

        response = self.client.get(
            reverse("planificator:istoric_detail", kwargs={"generation_id": generation.pk})
        )

        self.assertEqual(response.status_code, 404)

    def test_export_reads_owned_server_side_generation(self):
        generation = generation_for(self.user)
        response = self.client.post(
            reverse("planificator:generator_perioade_export"), {"generation_id": generation.pk}
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("Program", load_workbook(io.BytesIO(response.content)).sheetnames)

    def test_export_download_is_unchanged_for_htmx_header(self):
        generation = generation_for(self.user)
        response = self.client.post(
            reverse("planificator:generator_perioade_export"),
            {"generation_id": generation.pk},
            HTTP_HX_REQUEST="true",
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn('filename="program_cursuri_2026.xlsx"', response["Content-Disposition"])
        self.assertIn("Program", load_workbook(io.BytesIO(response.content)).sheetnames)

    def test_other_users_and_expired_generations_are_not_exportable(self):
        other = get_user_model().objects.create_user(username="other", password="x")
        other.user_permissions.add(self.permission)
        foreign = generation_for(other)
        expired = generation_for(self.user, expired=True)
        for generation in (foreign, expired):
            response = self.client.post(
                reverse("planificator:generator_perioade_export"), {"generation_id": generation.pk}
            )
            self.assertEqual(response.status_code, 404)

    def test_invalid_month_is_a_form_error_not_server_error(self):
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"input_file": csv_upload(), "year": 2026, "months": ["bogus"], "randomness": 5, "holidays": ""},
        )
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "Select a valid choice", status_code=400)

    def test_invalid_upload_extension_is_a_form_error(self):
        upload = SimpleUploadedFile("courses.xls", b"not-an-xls", content_type="application/octet-stream")
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {"input_file": upload, "year": 2026, "months": ["1"], "randomness": 5, "holidays": ""},
        )
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, ".csv", status_code=400)

    def test_invalid_year_randomness_and_holiday_markup_are_form_errors(self):
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {
                "input_file": csv_upload(),
                "year": 1900,
                "months": ["1"],
                "randomness": 11,
                "holidays": '<img src=x onerror="alert(1)">',
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertNotContains(response, '<img src=x onerror="alert(1)">', status_code=400)

    def test_corrupt_saved_settings_are_safely_normalized(self):
        AppSetting.objects.create(
            user=self.user,
            scope="schedule_generator",
            payload={"year": "bad", "months": ["bad", 13], "randomness": 99, "holidays": ["not-a-date"]},
        )
        response = self.client.get(reverse("planificator:generator_perioade"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_months"], [])

    def test_incomplete_schedule_is_not_persisted(self):
        holidays = "\n".join(
            f"{day:02d}.01.2026" for day in range(1, 32)
            if datetime(2026, 1, day).weekday() < 5
        )
        response = self.client.post(
            reverse("planificator:generator_perioade"),
            {
                "input_file": csv_upload(rows=["Blocked,2 days,https://example.com/blocked,100"]),
                "year": 2026,
                "months": ["1"],
                "randomness": 5,
                "holidays": holidays,
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertContains(response, "Program incomplet", status_code=400)
        self.assertFalse(ScheduleGeneration.objects.exists())

    def test_sample_csv_is_permission_protected_and_downloadable(self):
        response = self.client.get(reverse("planificator:generator_perioade_sample_csv"))
        self.assertEqual(response.status_code, 200)
        self.assertIn('filename="model_cursuri.csv"', response["Content-Disposition"])


class SettingsStoreTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.first = get_user_model().objects.create_user(username="first")
        cls.second = get_user_model().objects.create_user(username="second")
        AppSetting.objects.create(scope="schedule_generator", payload={"year": 2027, "randomness": 2})

    def test_global_defaults_and_user_overrides_are_merged(self):
        save_settings("schedule_generator", self.first, {"year": 2028, "months": [1]})
        self.assertEqual(get_settings("schedule_generator", self.first)["year"], 2028)
        self.assertEqual(get_settings("schedule_generator", self.first)["randomness"], 2)
        self.assertEqual(get_settings("schedule_generator", self.second)["year"], 2027)

    def test_users_cannot_overwrite_each_other(self):
        save_settings("schedule_generator", self.first, {"months": [1]})
        save_settings("schedule_generator", self.second, {"months": [2]})
        self.assertEqual(get_settings("schedule_generator", self.first)["months"], [1])
        self.assertEqual(get_settings("schedule_generator", self.second)["months"], [2])

    def test_database_errors_are_not_silenced(self):
        with self.assertLogs("apps.planificator.settings_store", level="ERROR"):
            with patch("apps.planificator.settings_store.AppSetting.objects.filter", side_effect=DatabaseError("down")):
                with self.assertRaises(DatabaseError):
                    get_settings("schedule_generator", self.first)


class FileHandlerTests(TestCase):
    def test_at_delimited_csv_returns_typed_rows(self):
        content = (
            'Title@"Durata Curs"@investitie@Permalink\n'
            '"Curs real unu"@"1 zi"@"200 euro"@https://example.com/unu/\n'
            '"Curs real doi"@"2 zile"@"300 euro"@https://example.com/doi/\n'
        ).encode()
        rows = read_input_file(content, ".csv")
        self.assertEqual(rows[0].title, "Curs real unu")
        self.assertEqual(rows[1].duration, 2)

    def test_xlsx_is_read_in_read_only_compatible_shape(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Title", "Durata Curs", "Permalink", "investitie"])
        sheet.append(["Curs", "3 zile", "https://example.com/curs", None])
        output = io.BytesIO()
        workbook.save(output)
        rows = read_input_file(output.getvalue(), ".xlsx")
        self.assertEqual(rows[0].investment, "")

    def test_blank_required_value_has_row_specific_error(self):
        with self.assertRaisesRegex(Exception, "Rândul 2: Permalink"):
            read_input_file(b"Title,Durata Curs,Permalink\nCurs,2 zile,\n", ".csv")

    def test_excel_export_preserves_duplicate_titles_and_neutralizes_formulas(self):
        schedule = []
        for source_row, investment in ((2, "+2"), (3, "@cmd")):
            schedule.append({
                "source_row": source_row,
                "original_order": source_row - 2,
                "Title": "=1+1",
                "Permalink": "https://example.com",
                "Durata Curs": "-3 zile",
                "investitie": investment,
                "date_range": "05.01.2026",
                "month": 1,
            })
        workbook = load_workbook(io.BytesIO(create_excel_export(schedule, 2026)))
        sheet = workbook["Program"]
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet["A2"].data_type, "s")
        self.assertEqual(sheet["A2"].value, "'=1+1")
        self.assertEqual(sheet["D2"].value, "'+2")
        self.assertEqual(sheet["E1"].value, "Ianuarie")


class GenerationCleanupTests(TestCase):
    def test_cleanup_command_removes_only_expired_generations(self):
        user = get_user_model().objects.create_user(username="cleanup")
        expired = generation_for(user, expired=True)
        active = generation_for(user)
        call_command("purge_expired_schedule_generations", verbosity=0)
        self.assertFalse(ScheduleGeneration.objects.filter(pk=expired.pk).exists())
        self.assertTrue(ScheduleGeneration.objects.filter(pk=active.pk).exists())
