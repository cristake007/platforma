import csv
import io
import re
import unicodedata
import uuid
from copy import deepcopy
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from zipfile import BadZipFile, ZIP_DEFLATED, ZipFile

from django.core.exceptions import ValidationError
from django.core.files.base import ContentFile
from django.db import transaction
from django.http import Http404, HttpResponse
from django.utils import timezone

from apps.media_library.services import (
    require_owned_layout_assets,
    serialize_media_assets,
)
from django.utils.text import slugify
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import (
    DiplomaGenerationBatch,
    DiplomaTemplate,
    GeneratedDiploma,
    Participant,
    ParticipantImportDraft,
    ParticipantList,
)
from .pdf_renderer import render_diploma_pdf
from .selectors import (
    get_owned_generation_batch,
    get_owned_generated_diploma,
    get_owned_import_draft_for_update,
    get_owned_participant,
    get_owned_participant_for_update,
    get_owned_participant_list,
    get_owned_participant_list_for_update,
    get_owned_template,
    get_owned_template_for_update,
    list_generated_diplomas_for_batch,
    list_owned_generation_batches,
)
from .validators import MAX_PARTICIPANT_ROWS, validate_layout_json


SAMPLE_PARTICIPANT = {
    "full_name": "Andrei Popescu",
    "date_of_birth": "12.04.1990",
    "place_of_birth": "Brașov",
    "certificate_number": "TTK-2026-001",
}

PARTICIPANT_DATE_ERROR = "Data nașterii trebuie să fie în format DD.MM.YYYY."
PARTICIPANT_COLUMNS = {
    "full_name": {"full_name", "nume complet", "nume si prenume", "nume și prenume"},
    "date_of_birth": {"date_of_birth", "data nasterii", "data nașterii"},
    "place_of_birth": {"place_of_birth", "locul nasterii", "locul nașterii"},
    "certificate_number": {
        "certificate_number",
        "numar certificat",
        "număr certificat",
    },
}


def _typography(*, size: int, color: str = "#164194", bold: bool = False, align: str = "center", font: str = "Lora") -> dict:
    return {
        "fontFamily": font,
        "fontSize": size,
        "bold": bold,
        "italic": False,
        "underline": False,
        "color": color,
        "align": align,
        "lineHeight": 1.18,
        "letterSpacing": 0,
        "textTransform": "none",
    }


def _base_element(*, element_id: str, element_type: str, label: str, x_mm: int, y_mm: int, width_mm: int, height_mm: int, z_index: int, style: dict) -> dict:
    return {
        "id": element_id,
        "type": element_type,
        "label": label,
        "x_mm": x_mm,
        "y_mm": y_mm,
        "width_mm": width_mm,
        "height_mm": height_mm,
        "rotation": 0,
        "zIndex": z_index,
        "locked": False,
        "visible": True,
        "style": style,
    }


def build_blank_layout(*, page_size: str = "A4", orientation: str = "landscape") -> dict:
    width_mm, height_mm = ((297, 210) if orientation == "landscape" else (210, 297))
    return validate_layout_json({
        "version": 2,
        "page": {
            "size": page_size,
            "orientation": orientation,
            "width_mm": width_mm,
            "height_mm": height_mm,
            "grid_mm": 1,
            "major_grid_mm": 10,
            "background": None,
        },
        "elements": [],
    })


def build_default_layout(*, page_size: str = "A4", orientation: str = "landscape") -> dict:
    width_mm, height_mm = ((297, 210) if orientation == "landscape" else (210, 297))

    def sx(value: int) -> int:
        return max(1, round(value * width_mm / 297))

    def sy(value: int) -> int:
        return max(1, round(value * height_mm / 210))

    elements = [
        {
            **_base_element(element_id="title", element_type="text", label="Titlu", x_mm=sx(53), y_mm=sy(28), width_mm=sx(191), height_mm=sy(19), z_index=10, style=_typography(size=46, bold=True)),
            "text": "DIPLOMĂ DE ABSOLVIRE",
        },
        {
            **_base_element(element_id="intro", element_type="text", label="Introducere", x_mm=sx(85), y_mm=sy(54), width_mm=sx(127), height_mm=sy(9), z_index=20, style=_typography(size=16, color="#304253", font="Inter")),
            "text": "Se acordă domnului / doamnei",
        },
        {
            **_base_element(element_id="full_name", element_type="variable", label="Nume complet", x_mm=sx(69), y_mm=sy(66), width_mm=sx(159), height_mm=sy(18), z_index=30, style=_typography(size=38, bold=True)),
            "variable": "full_name",
            "placeholder": "Nume complet",
        },
        {
            **_base_element(element_id="date_of_birth", element_type="variable", label="Data nașterii", x_mm=sx(66), y_mm=sy(91), width_mm=sx(66), height_mm=sy(10), z_index=40, style=_typography(size=16, color="#304253", font="Inter")),
            "variable": "date_of_birth",
            "placeholder": "Data nașterii",
        },
        {
            **_base_element(element_id="place_of_birth", element_type="variable", label="Locul nașterii", x_mm=sx(165), y_mm=sy(91), width_mm=sx(66), height_mm=sy(10), z_index=50, style=_typography(size=16, color="#304253", font="Inter")),
            "variable": "place_of_birth",
            "placeholder": "Locul nașterii",
        },
        {
            **_base_element(element_id="certificate_number", element_type="variable", label="Număr certificat", x_mm=sx(106), y_mm=sy(106), width_mm=sx(85), height_mm=sy(10), z_index=60, style=_typography(size=16, color="#304253", bold=True, font="Inter")),
            "variable": "certificate_number",
            "placeholder": "Număr certificat",
        },
        {
            **_base_element(
                element_id="course_table",
                element_type="table",
                label="Tabel curs",
                x_mm=sx(48),
                y_mm=sy(128),
                width_mm=sx(201),
                height_mm=sy(29),
                z_index=70,
                style={
                    "fontFamily": "Inter",
                    "fontSize": 14,
                    "bold": False,
                    "color": "#304253",
                    "align": "center",
                    "borderColor": "#164194",
                    "headerBackground": "#edf3f9",
                },
            ),
            "columns": ["Denumirea cursului", "Cod curs", "Durată"],
            "rows": [["Inspector SSM", "SSM-01", "80 ore"]],
        },
        {
            **_base_element(element_id="signature_left", element_type="text", label="Semnătură director", x_mm=sx(42), y_mm=sy(173), width_mm=sx(69), height_mm=sy(11), z_index=80, style=_typography(size=14, color="#304253", bold=True, font="Inter")),
            "text": "DIRECTOR GENERAL",
        },
        {
            **_base_element(element_id="signature_right", element_type="text", label="Semnătură responsabil", x_mm=sx(186), y_mm=sy(173), width_mm=sx(69), height_mm=sy(11), z_index=90, style=_typography(size=14, color="#304253", bold=True, font="Inter")),
            "text": "RESPONSABIL CURS",
        },
    ]
    return validate_layout_json({
        "version": 2,
        "page": {
            "size": page_size,
            "orientation": orientation,
            "width_mm": width_mm,
            "height_mm": height_mm,
            "grid_mm": 1,
            "major_grid_mm": 10,
            "background": None,
        },
        "elements": elements,
    })


def create_diploma_template(*, owner, data: dict) -> DiplomaTemplate:
    layout = build_blank_layout(page_size=data["page_size"], orientation=data["orientation"])
    template = DiplomaTemplate(
        owner=owner,
        name=data["name"],
        category=data.get("category", DiplomaTemplate.DEFAULT_CATEGORY),
        description=data.get("description", ""),
        page_size=data["page_size"],
        orientation=data["orientation"],
        page_width_mm=layout["page"]["width_mm"],
        page_height_mm=layout["page"]["height_mm"],
        grid_size_mm=layout["page"]["grid_mm"],
        major_grid_size_mm=layout["page"]["major_grid_mm"],
        layout_json=layout,
    )
    template.full_clean()
    template.save()
    return template


def update_diploma_template_layout(*, owner, template_id, layout: dict) -> DiplomaTemplate:
    normalized = validate_layout_json(layout)
    with transaction.atomic():
        template = get_owned_template_for_update(user=owner, template_id=template_id)
        require_owned_layout_assets(owner=owner, layout=normalized, for_update=True)
        template.layout_json = normalized
        template.full_clean()
        template.save(update_fields=("layout_json", "updated_at"))
    return template


def delete_diploma_template(*, owner, template_id) -> None:
    with transaction.atomic():
        template = get_owned_template_for_update(user=owner, template_id=template_id)
        template.delete()


def build_preview_data() -> dict:
    return deepcopy(SAMPLE_PARTICIPANT)


def _normalize_header(value) -> str:
    text = " ".join(str(value or "").strip().lower().replace("_", " ").split())
    return "".join(
        character
        for character in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(character)
    )


def _suggest_participant_mapping(headers) -> dict[str, int]:
    aliases = {
        field: {_normalize_header(alias) for alias in values}
        for field, values in PARTICIPANT_COLUMNS.items()
    }
    suggestions = {}
    for field, accepted in aliases.items():
        matches = [
            index
            for index, header in enumerate(headers)
            if _normalize_header(header) in accepted
        ]
        if len(matches) == 1:
            suggestions[field] = matches[0]
    return suggestions


def _strict_date(value) -> tuple[date, str]:
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        return value, value.strftime("%d.%m.%Y")

    text = str(value or "").strip()
    if not re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text):
        raise ValidationError(PARTICIPANT_DATE_ERROR)
    try:
        parsed = datetime.strptime(text, "%d.%m.%Y").date()
    except ValueError as exc:
        raise ValidationError(PARTICIPANT_DATE_ERROR) from exc
    if parsed.strftime("%d.%m.%Y") != text:
        raise ValidationError(PARTICIPANT_DATE_ERROR)
    return parsed, text


def _display_cell(value) -> str:
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    return str(value or "").strip()


def _validate_participant_row(*, values, column_mapping, source_row):
    raw = {
        field: values[index] if index < len(values) else None
        for field, index in column_mapping.items()
    }
    display = {field: _display_cell(value) for field, value in raw.items()}
    errors = []

    limits = {
        "full_name": (200, "Numele complet"),
        "place_of_birth": (200, "Locul nașterii"),
        "certificate_number": (100, "Numărul certificatului"),
    }
    for field, (max_length, label) in limits.items():
        if not display[field]:
            errors.append(f"{label} este obligatoriu.")
        elif len(display[field]) > max_length:
            errors.append(f"{label} poate avea cel mult {max_length} de caractere.")

    try:
        _, normalized_date = _strict_date(raw["date_of_birth"])
        display["date_of_birth"] = normalized_date
    except ValidationError:
        errors.append(PARTICIPANT_DATE_ERROR)

    row = {"source_row": source_row, **display}
    if errors:
        return None, {**row, "errors": errors}
    return row, None


def _parse_csv_table(content: bytes):
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise ValidationError("Fișierul CSV trebuie să fie codificat UTF-8.") from exc
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";"
    rows = list(csv.reader(io.StringIO(text, newline=""), delimiter=delimiter))
    if not rows:
        raise ValidationError("Fișierul nu conține niciun rând.")
    return [
        (
            row_number,
            values,
            [{"is_excel_date": False, "number_format": ""} for _ in values],
        )
        for row_number, values in enumerate(rows, start=1)
    ]


def _parse_xlsx_tables(content: bytes, *, first_row_has_headers: bool):
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
        raise ValidationError("Fișierul XLSX nu poate fi citit.") from exc
    try:
        worksheet_tables = []
        for worksheet in workbook.worksheets:
            if worksheet.sheet_state != "visible":
                continue
            parsed_rows = []
            for row_number, cells in enumerate(worksheet.iter_rows(), start=1):
                values = [cell.value for cell in cells]
                metadata = [
                    {
                        "is_excel_date": isinstance(cell.value, (date, datetime)),
                        "number_format": str(cell.number_format or ""),
                    }
                    for cell in cells
                ]
                while values and not _display_cell(values[-1]):
                    values.pop()
                    metadata.pop()
                if any(_display_cell(value) for value in values):
                    parsed_rows.append((row_number, values, metadata))
            has_enough_rows = not first_row_has_headers or len(parsed_rows) > 1
            if (
                parsed_rows
                and has_enough_rows
                and max(len(values) for _, values, _ in parsed_rows)
                >= len(PARTICIPANT_COLUMNS)
            ):
                worksheet_tables.append((worksheet.title, parsed_rows))

        if not worksheet_tables:
            raise ValidationError(
                "Fișierul XLSX nu conține nicio foaie cu cel puțin patru coloane."
            )

        return worksheet_tables
    finally:
        workbook.close()


def _participant_upload_tables(*, upload, first_row_has_headers: bool):
    content = upload.read()
    extension = Path(upload.name).suffix.lower()
    if extension == ".csv":
        return [("", _parse_csv_table(content))]
    elif extension == ".xlsx":
        return _parse_xlsx_tables(
            content,
            first_row_has_headers=first_row_has_headers,
        )
    raise ValidationError("Fișierul trebuie să fie în format CSV sau XLSX.")


def _discover_participant_table(*, table_rows, first_row_has_headers: bool) -> dict:
    table_rows = [
        row for row in table_rows if any(_display_cell(value) for value in row[1])
    ]
    if not table_rows:
        raise ValidationError("Fișierul nu conține niciun rând cu date.")

    header_values = table_rows[0][1] if first_row_has_headers else []
    data_rows = table_rows[1:] if first_row_has_headers else table_rows
    if not data_rows:
        raise ValidationError("Fișierul nu conține rânduri de participanți.")
    if len(data_rows) > MAX_PARTICIPANT_ROWS:
        raise ValidationError(
            f"Fișierul poate conține cel mult {MAX_PARTICIPANT_ROWS} de participanți."
        )

    column_count = max(len(values) for _, values, _ in table_rows)
    if column_count < len(PARTICIPANT_COLUMNS):
        raise ValidationError("Fișierul trebuie să conțină cel puțin patru coloane.")

    columns = []
    for index in range(column_count):
        label = _display_cell(header_values[index]) if index < len(header_values) else ""
        if not label:
            label = f"Coloana {index + 1}"
        samples = []
        for _, values, _ in data_rows:
            sample = _display_cell(values[index]) if index < len(values) else ""
            if sample and sample not in samples:
                samples.append(sample)
            if len(samples) == 2:
                break
        sample_text = ", ".join(samples)
        display_label = f"{label} — ex. {sample_text}" if sample_text else label
        columns.append(
            {
                "index": index,
                "label": label,
                "display_label": display_label,
                "samples": samples,
            }
        )

    source_rows = []
    for source_row, values, metadata in data_rows:
        source_rows.append(
            {
                "source_row": source_row,
                "values": [_display_cell(value) for value in values],
                "metadata": metadata,
            }
        )
    return {
        "columns": columns,
        "source_rows": source_rows,
        "suggested_mapping": (
            _suggest_participant_mapping(header_values)
            if first_row_has_headers
            else {}
        ),
    }


def discover_participant_upload(
    *,
    upload,
    first_row_has_headers: bool,
    worksheet_name: str | None = None,
) -> dict:
    tables = _participant_upload_tables(
        upload=upload,
        first_row_has_headers=first_row_has_headers,
    )
    if worksheet_name is not None:
        tables = [table for table in tables if table[0] == worksheet_name]
        if not tables:
            raise ValidationError("Foaia XLSX selectată nu este disponibilă.")
    elif len(tables) > 1:
        raise ValidationError("Selectează foaia XLSX care trebuie importată.")
    _, table_rows = tables[0]
    return _discover_participant_table(
        table_rows=table_rows,
        first_row_has_headers=first_row_has_headers,
    )


def validate_participant_mapping(*, source_rows, column_mapping) -> dict:
    expected_fields = set(PARTICIPANT_COLUMNS)
    if set(column_mapping) != expected_fields:
        raise ValidationError("Toate cele patru câmpuri trebuie asociate unei coloane.")
    try:
        normalized_mapping = {
            field: int(index) for field, index in column_mapping.items()
        }
    except (TypeError, ValueError) as exc:
        raise ValidationError("Asocierea coloanelor nu este validă.") from exc
    if len(set(normalized_mapping.values())) != len(normalized_mapping):
        raise ValidationError("Fiecare câmp trebuie asociat unei coloane diferite.")

    valid_rows = []
    invalid_rows = []
    warnings = []
    seen_certificates = set()
    for source in source_rows:
        values = source["values"]
        valid, invalid = _validate_participant_row(
            values=values,
            column_mapping=normalized_mapping,
            source_row=source["source_row"],
        )
        if invalid:
            invalid_rows.append(invalid)
            continue
        certificate = valid["certificate_number"].casefold()
        if certificate in seen_certificates:
            warnings.append(
                f"Rândul {source['source_row']}: numărul de certificat este duplicat în fișier."
            )
        seen_certificates.add(certificate)
        valid_rows.append(valid)
    return {
        "valid_rows": valid_rows,
        "invalid_rows": invalid_rows,
        "warnings": warnings,
    }


def parse_participant_upload(*, upload, worksheet_name: str | None = None) -> dict:
    discovered = discover_participant_upload(
        upload=upload,
        first_row_has_headers=True,
        worksheet_name=worksheet_name,
    )
    if set(discovered["suggested_mapping"]) != set(PARTICIPANT_COLUMNS):
        raise ValidationError("Coloanele trebuie asociate înainte de validarea participanților.")
    return validate_participant_mapping(
        source_rows=discovered["source_rows"],
        column_mapping=discovered["suggested_mapping"],
    )


def create_participant_import_draft(*, owner, data: dict) -> ParticipantImportDraft:
    tables = _participant_upload_tables(
        upload=data["source_file"],
        first_row_has_headers=data.get("first_row_has_headers", False),
    )
    discovered_sheets = [
        {
            "name": sheet_name,
            **_discover_participant_table(
                table_rows=table_rows,
                first_row_has_headers=data.get("first_row_has_headers", False),
            ),
        }
        for sheet_name, table_rows in tables
    ]
    discovered = discovered_sheets[0] if len(discovered_sheets) == 1 else None
    return ParticipantImportDraft.objects.create(
        owner=owner,
        list_name=data["list_name"],
        description=data.get("description", ""),
        course_name=data.get("course_name", ""),
        source_file_name=Path(data["source_file"].name).name[:255],
        source_sheets_json=discovered_sheets if len(discovered_sheets) > 1 else [],
        source_columns_json=discovered["columns"] if discovered else [],
        source_rows_json=discovered["source_rows"] if discovered else [],
        column_mapping_json=discovered["suggested_mapping"] if discovered else {},
        expires_at=timezone.now() + timedelta(hours=24),
    )


def select_participant_import_sheet(
    *, owner, draft_id, sheet_index: int
) -> ParticipantImportDraft:
    with transaction.atomic():
        draft = get_owned_import_draft_for_update(user=owner, draft_id=draft_id)
        if sheet_index < 0:
            raise ValidationError("Foaia XLSX selectată nu este disponibilă.")
        try:
            selected = draft.source_sheets_json[sheet_index]
        except (IndexError, TypeError):
            raise ValidationError("Foaia XLSX selectată nu este disponibilă.")
        draft.source_columns_json = selected["columns"]
        draft.source_rows_json = selected["source_rows"]
        draft.column_mapping_json = selected["suggested_mapping"]
        draft.save(
            update_fields=(
                "source_columns_json",
                "source_rows_json",
                "column_mapping_json",
            )
        )
        return draft


def apply_participant_import_mapping(*, owner, draft_id, column_mapping) -> ParticipantImportDraft:
    with transaction.atomic():
        draft = get_owned_import_draft_for_update(user=owner, draft_id=draft_id)
        available_indexes = {
            int(column["index"]) for column in draft.source_columns_json
        }
        try:
            requested_indexes = {int(index) for index in column_mapping.values()}
        except (TypeError, ValueError) as exc:
            raise ValidationError("Asocierea coloanelor nu este validă.") from exc
        if not requested_indexes.issubset(available_indexes):
            raise ValidationError("Asocierea conține o coloană inexistentă.")
        parsed = validate_participant_mapping(
            source_rows=draft.source_rows_json,
            column_mapping=column_mapping,
        )
        draft.column_mapping_json = {
            field: int(index) for field, index in column_mapping.items()
        }
        draft.mapping_confirmed = True
        draft.valid_rows_json = parsed["valid_rows"]
        draft.invalid_rows_json = parsed["invalid_rows"]
        draft.warnings_json = parsed["warnings"]
        draft.save(
            update_fields=(
                "column_mapping_json",
                "mapping_confirmed",
                "valid_rows_json",
                "invalid_rows_json",
                "warnings_json",
            )
        )
        return draft


def confirm_participant_import(*, owner, draft_id) -> ParticipantList:
    with transaction.atomic():
        draft = get_owned_import_draft_for_update(user=owner, draft_id=draft_id)
        if not draft.mapping_confirmed:
            raise ValidationError("Coloanele trebuie asociate înainte de confirmarea importului.")
        if not draft.valid_rows_json:
            raise ValidationError("Importul nu conține niciun rând valid.")
        participant_list = ParticipantList.objects.create(
            owner=owner,
            name=draft.list_name,
            description=draft.description,
            course_name=draft.course_name,
            source_file_name=draft.source_file_name,
            participant_count=len(draft.valid_rows_json),
        )
        participants = []
        for row in draft.valid_rows_json:
            parsed_date, _ = _strict_date(row["date_of_birth"])
            participants.append(
                Participant(
                    owner=owner,
                    participant_list=participant_list,
                    full_name=row["full_name"],
                    date_of_birth=parsed_date,
                    place_of_birth=row["place_of_birth"],
                    certificate_number=row["certificate_number"],
                    source_row=row["source_row"],
                )
            )
        Participant.objects.bulk_create(participants)
        draft.delete()
        return participant_list


def delete_participant_list(*, owner, participant_list_id) -> None:
    with transaction.atomic():
        participant_list = get_owned_participant_list_for_update(
            user=owner,
            participant_list_id=participant_list_id,
        )
        participant_list.delete()


def build_diploma_preview_context(
    *,
    owner,
    participant_list_id,
    participant_id,
    template_id,
) -> dict:
    try:
        participant_list_id = uuid.UUID(str(participant_list_id))
        participant_id = uuid.UUID(str(participant_id))
        template_id = uuid.UUID(str(template_id))
    except (TypeError, ValueError, AttributeError) as exc:
        raise Http404("Selecția pentru previzualizare nu este validă.") from exc
    participant_list = get_owned_participant_list(
        user=owner,
        participant_list_id=participant_list_id,
    )
    participant = get_owned_participant(user=owner, participant_id=participant_id)
    template = get_owned_template(user=owner, template_id=template_id)
    if participant.participant_list_id != participant_list.pk:
        raise Http404("Participantul nu aparține listei selectate.")
    return {
        "participant_list": participant_list,
        "participant": participant,
        "diploma_template": template,
        "layout": validate_layout_json(template.layout_json),
        "media_assets": serialize_media_assets(
            require_owned_layout_assets(owner=owner, layout=template.layout_json).values()
        ),
        "participant_data": {
            "full_name": participant.full_name,
            "date_of_birth": participant.date_of_birth.strftime("%d.%m.%Y"),
            "place_of_birth": participant.place_of_birth,
            "certificate_number": participant.certificate_number,
        },
    }


def generate_single_diploma(
    *,
    owner,
    participant_list_id,
    participant_id,
    template_id,
) -> GeneratedDiploma:
    stored_file_name = ""
    storage = None
    try:
        with transaction.atomic():
            participant_list = get_owned_participant_list_for_update(
                user=owner,
                participant_list_id=participant_list_id,
            )
            participant = get_owned_participant_for_update(
                user=owner,
                participant_id=participant_id,
            )
            template = get_owned_template_for_update(
                user=owner,
                template_id=template_id,
            )
            if participant.participant_list_id != participant_list.pk:
                raise Http404("Participantul nu aparține listei selectate.")

            pdf_bytes = render_diploma_pdf(template=template, participant=participant)
            generated = GeneratedDiploma(
                owner=owner,
                participant_list=participant_list,
                participant=participant,
                template=template,
                certificate_number=participant.certificate_number,
                participant_name=participant.full_name,
                participant_list_name=participant_list.name,
                template_name=template.name,
            )
            generated.pdf_file.save(
                "diploma.pdf",
                ContentFile(pdf_bytes),
                save=False,
            )
            storage = generated.pdf_file.storage
            stored_file_name = generated.pdf_file.name
            generated.full_clean()
            generated.save()
            return generated
    except Exception:
        if storage is not None and stored_file_name:
            storage.delete(stored_file_name)
        raise


def get_generated_diploma_download(*, owner, generated_diploma_id) -> GeneratedDiploma:
    return get_owned_generated_diploma(
        user=owner,
        generated_diploma_id=generated_diploma_id,
    )


def build_generated_diploma_filename(generated_diploma: GeneratedDiploma) -> str:
    certificate = _safe_filename_part(
        generated_diploma.certificate_number,
        fallback="certificat",
    )
    participant = _safe_filename_part(
        generated_diploma.participant_name,
        fallback="participant",
    )
    safe_suffix = "_".join(filter(None, (certificate, participant))) or "diploma"
    return f"diploma_{safe_suffix[:180]}.pdf"


def _safe_filename_part(value, *, fallback: str) -> str:
    separated = re.sub(r"[^\w]+", "-", str(value or ""), flags=re.UNICODE)
    normalized = slugify(separated, allow_unicode=False).replace("-", "_")
    normalized = re.sub(r"[^a-zA-Z0-9_]", "_", normalized)
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or fallback


def _batch_pdf_filename(participant: Participant) -> str:
    certificate = _safe_filename_part(
        participant.certificate_number,
        fallback="certificat",
    )
    participant_name = _safe_filename_part(
        participant.full_name,
        fallback="participant",
    )
    return f"{certificate[:80]}_{participant_name[:120]}.pdf"


def _available_batch_pdf_filename(*, batch, participant, storage) -> str:
    original = _batch_pdf_filename(participant)
    stem = original.removesuffix(".pdf")
    candidate = original
    suffix = 2
    while storage.exists(f"{batch.output_folder}/{candidate}"):
        candidate = f"{stem[:190]}_{suffix}.pdf"
        suffix += 1
    return candidate


def create_generation_batch(owner, participant_list_id, template_id):
    with transaction.atomic():
        participant_list = get_owned_participant_list_for_update(
            user=owner,
            participant_list_id=participant_list_id,
        )
        template = get_owned_template_for_update(
            user=owner,
            template_id=template_id,
        )
        normalized_layout = validate_layout_json(template.layout_json)
        require_owned_layout_assets(owner=owner, layout=normalized_layout)
        total_count = participant_list.participants.filter(owner=owner).count()
        if total_count == 0:
            raise ValidationError("Lista selectată nu conține participanți.")

        batch = DiplomaGenerationBatch(
            owner=owner,
            participant_list=participant_list,
            template=template,
            participant_list_name=participant_list.name,
            template_name=template.name,
            total_count=total_count,
        )
        batch.output_folder = (
            f"diplomas/{owner.pk}/{participant_list.pk}/{batch.pk}"
        )
        batch.full_clean()
        batch.save()
        return batch


def generate_batch_participant_pdf(batch, participant):
    if participant.owner_id != batch.owner_id:
        raise ValidationError("Participantul nu aparține utilizatorului lotului.")
    if participant.participant_list_id != batch.participant_list_id:
        raise ValidationError("Participantul nu aparține listei lotului.")

    pdf_bytes = render_diploma_pdf(
        template=batch.template,
        participant=participant,
    )
    generated = GeneratedDiploma(
        owner_id=batch.owner_id,
        participant_list=batch.participant_list,
        participant=participant,
        template=batch.template,
        batch=batch,
        certificate_number=participant.certificate_number,
        participant_name=participant.full_name,
        participant_list_name=batch.participant_list_display_name,
        template_name=batch.template_display_name,
    )
    storage = generated.pdf_file.storage
    stored_file_name = ""
    try:
        filename = _available_batch_pdf_filename(
            batch=batch,
            participant=participant,
            storage=storage,
        )
        generated.pdf_file.save(filename, ContentFile(pdf_bytes), save=False)
        stored_file_name = generated.pdf_file.name
        generated.full_clean()
        generated.save()
        return generated
    except Exception:
        if stored_file_name:
            storage.delete(stored_file_name)
        raise


def _participant_batch_error(participant) -> dict:
    return {
        "participant_id": str(participant.pk),
        "participant_name": participant.full_name,
        "certificate_number": participant.certificate_number,
        "message": "PDF-ul nu a putut fi generat pentru acest participant.",
    }


def finalize_generation_batch(batch):
    success_count = batch.generated_diplomas.count()
    failed_count = max(batch.total_count - success_count, 0)
    if success_count == batch.total_count:
        status = DiplomaGenerationBatch.Status.COMPLETED
    elif success_count:
        status = DiplomaGenerationBatch.Status.COMPLETED_WITH_ERRORS
    else:
        status = DiplomaGenerationBatch.Status.FAILED

    batch.status = status
    batch.success_count = success_count
    batch.failed_count = failed_count
    batch.completed_at = timezone.now()
    batch.save(
        update_fields=(
            "status",
            "success_count",
            "failed_count",
            "error_summary",
            "completed_at",
            "updated_at",
        )
    )
    return batch


def run_generation_batch(batch_id):
    with transaction.atomic():
        batch = (
            DiplomaGenerationBatch.objects.select_for_update(of=("self",))
            .select_related(
                "owner",
                "participant_list",
                "template",
            )
            .get(pk=batch_id)
        )
        if batch.status in {
            DiplomaGenerationBatch.Status.COMPLETED,
            DiplomaGenerationBatch.Status.COMPLETED_WITH_ERRORS,
            DiplomaGenerationBatch.Status.FAILED,
            DiplomaGenerationBatch.Status.RUNNING,
        }:
            return batch
        batch.status = DiplomaGenerationBatch.Status.RUNNING
        batch.started_at = timezone.now()
        batch.success_count = 0
        batch.failed_count = 0
        batch.error_summary = []
        batch.save(
            update_fields=(
                "status",
                "started_at",
                "success_count",
                "failed_count",
                "error_summary",
                "updated_at",
            )
        )

    participants = Participant.objects.filter(
        owner_id=batch.owner_id,
        participant_list_id=batch.participant_list_id,
    ).order_by("source_row", "full_name")
    errors = []
    success_count = 0
    attempted_count = 0
    try:
        for participant in participants.iterator():
            attempted_count += 1
            try:
                generate_batch_participant_pdf(batch, participant)
                success_count += 1
            except Exception:
                errors.append(_participant_batch_error(participant))
            DiplomaGenerationBatch.objects.filter(pk=batch.pk).update(
                success_count=success_count,
                failed_count=attempted_count - success_count,
                error_summary=errors,
                updated_at=timezone.now(),
            )
    except Exception:
        errors.append(
            {
                "message": "Generarea lotului a fost întreruptă înainte de finalizare."
            }
        )

    batch.refresh_from_db()
    batch.error_summary = errors
    return finalize_generation_batch(batch)


def _mark_pending_generation_batch_failed(batch_id):
    with transaction.atomic():
        batch = (
            DiplomaGenerationBatch.objects.select_for_update(of=("self",))
            .filter(
                pk=batch_id,
                status=DiplomaGenerationBatch.Status.PENDING,
            )
            .first()
        )
        if batch is None:
            return
        batch.status = DiplomaGenerationBatch.Status.FAILED
        batch.failed_count = batch.total_count
        batch.error_summary = [
            {"message": "Generarea lotului nu a putut porni."}
        ]
        batch.completed_at = timezone.now()
        batch.save(
            update_fields=(
                "status",
                "failed_count",
                "error_summary",
                "completed_at",
                "updated_at",
            )
        )


def _run_pending_generation_batch(batch_id):
    try:
        return run_generation_batch(batch_id)
    except Exception:
        _mark_pending_generation_batch_failed(batch_id)
        raise


def generate_diploma_batch(owner, participant_list_id, template_id):
    batch = create_generation_batch(owner, participant_list_id, template_id)
    return _run_pending_generation_batch(batch.pk)


def resume_generation_batch(owner, batch_id):
    batch = get_owned_generation_batch(owner, batch_id)
    if batch.status != DiplomaGenerationBatch.Status.PENDING:
        raise ValidationError(
            "Doar loturile aflate în așteptare pot fi reluate."
        )
    return _run_pending_generation_batch(batch.pk)


def build_batch_zip_response(owner, batch_id):
    batch = get_owned_generation_batch(owner, batch_id)
    generated_diplomas = list_generated_diplomas_for_batch(owner, batch.pk)
    archive_buffer = BytesIO()
    used_names = set()
    with ZipFile(archive_buffer, mode="w", compression=ZIP_DEFLATED) as archive:
        for generated in generated_diplomas:
            try:
                if not generated.pdf_file.storage.exists(generated.pdf_file.name):
                    continue
                with generated.pdf_file.open("rb") as pdf_file:
                    content = pdf_file.read()
            except (OSError, ValueError):
                continue

            archive_name = Path(generated.pdf_file.name).name
            archive_name = re.sub(r"[^a-zA-Z0-9_.-]", "_", archive_name)
            stem = Path(archive_name).stem or "diploma"
            suffix = Path(archive_name).suffix.lower() or ".pdf"
            unique_name = f"{stem}{suffix}"
            duplicate = 2
            while unique_name.casefold() in used_names:
                unique_name = f"{stem}_{duplicate}{suffix}"
                duplicate += 1
            used_names.add(unique_name.casefold())
            archive.writestr(unique_name, content)

    list_name = _safe_filename_part(
        batch.participant_list_display_name,
        fallback="lista",
    )
    archive_filename = (
        f"diplome_{list_name[:140]}_{batch.created_at:%Y-%m-%d}.zip"
    )
    response = HttpResponse(
        archive_buffer.getvalue(),
        content_type="application/zip",
    )
    response["Content-Disposition"] = f'attachment; filename="{archive_filename}"'
    return response


def build_generation_history_context(owner, filters):
    batches = list_owned_generation_batches(owner)
    participant_list = filters.get("participant_list")
    template = filters.get("template")
    status = filters.get("status")
    created_date = filters.get("date")
    if participant_list:
        batches = batches.filter(participant_list=participant_list)
    if template:
        batches = batches.filter(template=template)
    if status:
        batches = batches.filter(status=status)
    if created_date:
        batches = batches.filter(created_at__date=created_date)
    return {"batches": batches}
